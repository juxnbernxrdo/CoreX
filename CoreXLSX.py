import openpyxl 
from openpyxl.styles import Alignment
import os
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox

def procesar_archivo(archivo, tipo):
    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active

    # Centrar todas las celdas de la hoja
    for fila in hoja.iter_rows():
        for celda in fila:
            celda.alignment = Alignment(horizontal='center', vertical='center')

    # Centrar el encabezado
    for celda in hoja[1]:
        celda.alignment = Alignment(horizontal='center', vertical='center')

    # Definir las columnas según el tipo de archivo
    columnas_dict = {
        "Facturas": ['IDENTIFICACION PROVEEDOR (RUC/CI)', 'SERIE', 'SECUENCIAL', 'AUTORIZACION'],
        "Notas de Crédito": ['RUC', 'NC', 'AUTORIZACION', 'ESTABLECIMIENTO', 'PUNTO', 'SECUENCIAL', 'FACTURA APLICADA'],
        "Retenciones": ['RUC DEL AGENTE RETENCION', 'SERIE', 'SECUENCIA', 'CLAVE DE ACCESO (Comprobantes de Retencion Electronicos)']
    }

    columnas_a_procesar = columnas_dict.get(tipo, [])
    encabezados = [celda.value for celda in hoja[1]]
    indices_columnas = [encabezados.index(col) for col in columnas_a_procesar if col in encabezados]

    # Procesar celdas específicas
    for fila in hoja.iter_rows(min_row=2):
        for idx_col in indices_columnas:
            celda = fila[idx_col]
            if isinstance(celda.value, str):
                celda.value = celda.value.replace("'", "")

    # Guardar archivo modificado
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    nuevo_nombre = f"modificado_{tipo}_{timestamp}_{os.path.basename(archivo)}"
    archivo_modificado = os.path.join(os.path.dirname(archivo), nuevo_nombre)
    wb.save(archivo_modificado)

    messagebox.showinfo("Proceso finalizado", f"Archivo procesado y guardado como: {nuevo_nombre}")

def seleccionar_archivos(tipo):
    archivos = filedialog.askopenfilenames(title=f"Selecciona archivos {tipo}", filetypes=[("Archivos de Excel", "*.xlsx")])
    for archivo in archivos:
        procesar_archivo(archivo, tipo)

def main():
    root = tk.Tk()
    root.title("CoreXLSX")
    
    # Aumentamos el tamaño de la ventana
    root.geometry("500x600")  # Tamaño más grande
    root.config(bg="#F2F2F7")  # Fondo de color gris claro de macOS
    root.resizable(True, True)  # Habilitamos el redimensionamiento

    # Título con fuente moderna
    title_label = tk.Label(root, text="Selecciona el tipo de documento", font=("Helvetica Neue", 20), fg="#333333", bg="#F2F2F7")
    title_label.pack(pady=30)

    # Frame para los botones, con estilo minimalista
    frame = tk.Frame(root, bg="#F2F2F7")
    frame.pack(pady=20, fill=tk.BOTH, expand=True)

    # Estilo de los botones con un diseño limpio y sin bordes
    button_style = {
        'font': ("Helvetica Neue", 16),  # Aumentamos el tamaño de la fuente proporcionalmente
        'height': 3,
        'bd': 0,
        'padx': 50,
        'pady': 20,
        'relief': "flat",
        'bg': "#E1E4E8",  # Fondo gris macOS
        'fg': "#333333",  # Texto oscuro
        'activebackground': "#D1D5DB",  # Fondo al hacer clic
        'activeforeground': "#333333",  # Texto al hacer clic
        'width': 20,
        'anchor': "center"
    }

    # Botones con iconos y texto descriptivo
    btn_facturas = tk.Button(frame, text="📄 Facturas", command=lambda: seleccionar_archivos("Facturas"), **button_style)
    btn_nc = tk.Button(frame, text="🧾 Notas de Crédito", command=lambda: seleccionar_archivos("Notas de Crédito"), **button_style)
    btn_retenciones = tk.Button(frame, text="📑 Retenciones", command=lambda: seleccionar_archivos("Retenciones"), **button_style)

    # Colocamos los botones con espaciado vertical
    btn_facturas.pack(pady=15, fill=tk.X)
    btn_nc.pack(pady=15, fill=tk.X)
    btn_retenciones.pack(pady=15, fill=tk.X)

    # Ejecutar la ventana principal
    root.mainloop()

if __name__ == "__main__":
    main()