import openpyxl
from openpyxl.styles import Alignment
import os
import datetime

# Ruta donde se encuentran los archivos .xlsx
directorio = r''

# Función para procesar cada archivo .xlsx
def procesar_archivo(archivo):
    # Abrir el archivo Excel
    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active  # Seleccionar la hoja activa

    # Ajustar y centrar todos los valores en el documento
    for fila in hoja.iter_rows():
        for celda in fila:
            celda.alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar y centrar los valores de la primera fila
    for celda in hoja[1]:  # Primera fila
        celda.alignment = Alignment(horizontal='center', vertical='center')

    # Eliminar las comillas simples de las columnas específicas
    columnas_a_procesar = ['RUC', 'NC', 'AUTORIZACION', 'ESTABLECIMIENTO', 'PUNTO', 'SECUENCIAL', 'FACTURA APLICADA']

    # Obtener los índices de las columnas a partir de la primera fila (encabezado)
    encabezados = [celda.value for celda in hoja[1]]
    indices_columnas = [encabezados.index(columna) for columna in columnas_a_procesar if columna in encabezados]

    # Procesar las filas y eliminar las comillas simples en las columnas seleccionadas
    for fila in hoja.iter_rows(min_row=2):  # Empezar desde la segunda fila (datos)
        for idx_col in indices_columnas:
            celda = fila[idx_col]
            if isinstance(celda.value, str):
                celda.value = celda.value.replace("'", "")  # Eliminar comillas simples

    # Crear un nuevo nombre para el archivo modificado
    # Usamos la fecha y hora actual para crear un nombre único
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    nuevo_nombre = f"modificado_{timestamp}_{os.path.basename(archivo)}"

    # Guardar el archivo modificado con el nuevo nombre
    archivo_modificado = os.path.join(os.path.dirname(archivo), nuevo_nombre)
    wb.save(archivo_modificado)

    print(f'Archivo procesado y guardado como: {nuevo_nombre}')

# Procesar todos los archivos .xlsx en el directorio
for archivo_nombre in os.listdir(directorio):
    if archivo_nombre.endswith('.xlsx'):
        archivo_ruta = os.path.join(directorio, archivo_nombre)
        procesar_archivo(archivo_ruta)
