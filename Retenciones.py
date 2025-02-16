import openpyxl
from openpyxl.styles import Alignment
import os
import datetime

directorio = r''

def procesar_archivo(archivo):
    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active

    for fila in hoja.iter_rows():
        for celda in fila:
            celda.alignment = Alignment(horizontal='center', vertical='center')

    for celda in hoja[1]:
        celda.alignment = Alignment(horizontal='center', vertical='center')

    columnas_a_procesar = ['RUC DEL AGENTE RETENCION', 'SERIE', 'SECUENCIA', 'CLAVE DE ACCESO (Comprobantes de Retencion Electronicos)']
    encabezados = [celda.value for celda in hoja[1]]
    indices_columnas = []

    for columna in columnas_a_procesar:
        if columna in encabezados:
            indices_columnas.append(encabezados.index(columna))
        else:
            print(f"Columna '{columna}' no encontrada en los encabezados.")

    for fila in hoja.iter_rows(min_row=2):
        for idx_col in indices_columnas:
            celda = fila[idx_col]
            if isinstance(celda.value, str):
                celda.value = celda.value.replace("'","")

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    nuevo_nombre = f"modificado_{timestamp}_{os.path.basename(archivo)}"
    archivo_modificado = os.path.join(os.path.dirname(archivo), nuevo_nombre)
    wb.save(archivo_modificado)

    print(f'Archivo procesado y guardado como: {nuevo_nombre}')

for archivo_nombre in os.listdir(directorio):
    if archivo_nombre.endswith('.xlsx'):
        archivo_ruta = os.path.join(directorio, archivo_nombre)
        procesar_archivo(archivo_ruta)

