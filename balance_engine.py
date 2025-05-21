import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Optional, Dict, Callable

def calcular_balance(df: pd.DataFrame, callback: Optional[Callable] = None) -> pd.DataFrame:
    required_columns = ["categoria", "tipo", "valor"]
    if not all(col in df.columns for col in required_columns):
        raise ValueError("DataFrame must contain 'categoria', 'tipo', 'valor' columns.")

    detailed_df = df[required_columns].copy()
    detailed_df["valor"] = pd.to_numeric(detailed_df["valor"], errors="coerce").fillna(0)

    category_sums = detailed_df.groupby(["categoria", "tipo"])["valor"].sum().reset_index()

    activos_total = detailed_df[detailed_df["categoria"].str.contains("activos", case=False)]["valor"].sum()
    pasivos_total = detailed_df[detailed_df["categoria"].str.contains("pasivos", case=False)]["valor"].sum()
    patrimonio_total = detailed_df[detailed_df["categoria"].str.contains("patrimonio", case=False)]["valor"].sum()
    pasivos_patrimonio_total = pasivos_total + patrimonio_total

    summary_data = [
        {"categoria": "TOTALES", "tipo": "Total Activos", "valor": activos_total},
        {"categoria": "TOTALES", "tipo": "Total Pasivos", "valor": pasivos_total},
        {"categoria": "TOTALES", "tipo": "Total Patrimonio", "valor": patrimonio_total},
        {"categoria": "TOTALES", "tipo": "Total Pasivos + Patrimonio", "valor": pasivos_patrimonio_total}
    ]
    summary_df = pd.DataFrame(summary_data)

    result_df = pd.concat([category_sums, summary_df], ignore_index=True)

    if callback:
        callback(result_df)

    return result_df

def calcular_ratios(df_balance: pd.DataFrame) -> Dict[str, float]:
    try:
        activos_total = df_balance[df_balance["categoria"].str.contains("activos", case=False)]["valor"].sum()
        pasivos_total = df_balance[df_balance["categoria"].str.contains("pasivos", case=False)]["valor"].sum()
        patrimonio_total = df_balance[df_balance["categoria"].str.contains("patrimonio", case=False)]["valor"].sum()

        activos_corrientes = df_balance[
            (df_balance["categoria"].str.contains("corrientes", case=False)) &
            (df_balance["categoria"].str.contains("activos", case=False))
        ]["valor"].sum()
        pasivos_corrientes = df_balance[
            (df_balance["categoria"].str.contains("corrientes", case=False)) &
            (df_balance["categoria"].str.contains("pasivos", case=False))
        ]["valor"].sum()

        endeudamiento = pasivos_total / activos_total if activos_total > 0 else float('inf')
        liquidez = activos_corrientes / pasivos_corrientes if pasivos_corrientes > 0 else float('inf')
        solvencia = patrimonio_total / activos_total if activos_total > 0 else float('inf')

        return {
            "Endeudamiento": round(endeudamiento, 2) if endeudamiento != float('inf') else 0.0,
            "Liquidez": round(liquidez, 2) if liquidez != float('inf') else 0.0,
            "Solvencia": round(solvencia, 2) if solvencia != float('inf') else 0.0
        }
    except Exception as e:
        raise ValueError(f"Error calculating ratios: {str(e)}")

def generar_diagnostico(ratios: Dict[str, float], totales: Optional[Dict[str, float]] = None) -> str:
    partes = []
    if totales:
        partes.append(f'<span style="color: #333333;">📊 Total Activos: ${totales["Total Activos"]:,.2f}</span>')
        partes.append(f'<span style="color: #333333;">📊 Total Pasivos: ${totales["Total Pasivos"]:,.2f}</span>')
        partes.append(f'<span style="color: #333333;">📊 Total Patrimonio: ${totales["Total Patrimonio"]:,.2f}</span>')
        partes.append(f'<span style="color: #333333;">📊 Total Pasivos + Patrimonio: ${totales["Total Pasivos + Patrimonio"]:,.2f}</span>')
        partes.append("")

    if ratios['Endeudamiento'] < 0.5:
        partes.append('<span style="color: #28a745;">✅ Bajo endeudamiento (solidez financiera).</span>')
    elif ratios['Endeudamiento'] == 0.0:
        partes.append('<span style="color: #6c757d;">⚠️ Endeudamiento no calculable (falta de datos).</span>')
    else:
        partes.append('<span style="color: #dc3545;">⚠️ Endeudamiento alto (riesgo financiero).</span>')

    if ratios['Liquidez'] > 1:
        partes.append(f'<span style="color: #28a745;">✅ Liquidez adecuada ({ratios["Liquidez"]:.2f}).</span>')
    elif ratios['Liquidez'] == 0.0:
        partes.append('<span style="color: #6c757d;">⚠️ Liquidez no calculable (falta de datos).</span>')
    else:
        partes.append(f'<span style="color: #dc3545;">⚠️ Liquidez baja ({ratios["Liquidez"]:.2f}).</span>')

    if ratios['Solvencia'] > 0:
        partes.append(f'<span style="color: #28a745;">✅ Solvencia positiva ({ratios["Solvencia"]:.2f}).</span>')
    elif ratios['Solvencia'] == 0.0:
        partes.append('<span style="color: #6c757d;">⚠️ Solvencia no calculable (falta de datos).</span>')
    else:
        partes.append(f'<span style="color: #dc3545;">⚠️ Solvencia negativa ({ratios["Solvencia"]:.2f}).</span>')

    return "<br>".join(partes)

def exportar_balance_profesional(
    nombre_empresa: str,
    fecha_balance: str,
    df_balance: pd.DataFrame,
    archivo_salida: str,
    ratios: Optional[Dict[str, float]] = None,
    diagnostico: Optional[str] = None
) -> None:
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance General"

        bold_font = Font(bold=True, size=12)
        header_font = Font(bold=True, size=14)
        small_bold = Font(bold=True, size=11)
        center = Alignment(horizontal="center")
        right = Alignment(horizontal="right")
        border = Border(bottom=Side(border_style="thin"), top=Side(border_style="thin"),
                       left=Side(border_style="thin"), right=Side(border_style="thin"))

        ws.merge_cells("A1:C1")
        ws["A1"] = f"Balance General - {nombre_empresa}"
        ws["A1"].font = header_font
        ws["A1"].alignment = center

        ws.merge_cells("A2:C2")
        ws["A2"] = f"Al {fecha_balance}"
        ws["A2"].font = small_bold
        ws["A2"].alignment = center

        fila = 4
        current_category = ""

        ws[f"A{fila}"] = "Categoría"
        ws[f"B{fila}"] = "Tipo"
        ws[f"C{fila}"] = "Valor"
        for col in ["A", "B", "C"]:
            cell = ws[f"{col}{fila}"]
            cell.font = small_bold
            cell.border = border
            cell.alignment = center
        fila += 1

        for _, row in df_balance.iterrows():
            categoria = str(row["categoria"]).strip().title()
            tipo = str(row["tipo"]).strip().title()
            valor = float(row["valor"]) if pd.notna(row["valor"]) else 0.0

            if categoria and categoria != current_category:
                ws[f"A{fila}"] = categoria
                ws[f"A{fila}"].font = bold_font
                ws[f"A{fila}"].border = border
                current_category = categoria
                fila += 1

            ws[f"A{fila}"] = "" if categoria == current_category else categoria
            ws[f"B{fila}"] = tipo
            ws[f"C{fila}"] = valor
            ws[f"C{fila}"].number_format = '#,##0.00'
            ws[f"C{fila}"].alignment = right
            for col in ["A", "B", "C"]:
                ws[f"{col}{fila}"].border = border
            fila += 1

        if ratios:
            ws[f"A{fila}"] = "Ratios Financieros"
            ws[f"A{fila}"].font = bold_font
            ws[f"A{fila}"].border = border
            fila += 1
            for key, value in ratios.items():
                ws[f"A{fila}"] = key
                ws[f"C{fila}"] = value
                ws[f"C{fila}"].number_format = '0.00'
                ws[f"C{fila}"].alignment = right
                for col in ["A", "B", "C"]:
                    ws[f"{col}{fila}"].border = border
                fila += 1

        if diagnostico:
            ws[f"A{fila}"] = "Diagnóstico Financiero"
            ws[f"A{fila}"].font = bold_font
            ws[f"A{fila}"].border = border
            fila += 1
            for line in diagnostico.split("<br>"):
                cleaned_line = line.replace('<span style="color: #28a745;">', '').replace('<span style="color: #dc3545;">', '').replace('<span style="color: #6c757d;">', '').replace('<span style="color: #333333;">', '').replace('</span>', '')
                ws[f"A{fila}"] = cleaned_line
                for col in ["A", "B", "C"]:
                    ws[f"{col}{fila}"].border = border
                fila += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20

        if diagnostico:
            ws_diag = wb.create_sheet("Diagnóstico")
            ws_diag["A1"] = "Análisis Financiero"
            ws_diag["A1"].font = header_font
            ws_diag["A1"].alignment = Alignment(horizontal="left")
            for idx, line in enumerate(diagnostico.split("<br>"), start=3):
                cleaned_line = line.replace('<span style="color: #28a745;">', '').replace('<span style="color: #dc3545;">', '').replace('<span style="color: #6c757d;">', '').replace('<span style="color: #333333;">', '').replace('</span>', '')
                ws_diag[f"A{idx}"] = cleaned_line
            ws_diag.column_dimensions["A"].width = 60

        wb.save(archivo_salida)
    except PermissionError:
        raise ValueError(f"No se puede escribir en {archivo_salida}. Verifica los permisos.")
    except Exception as e:
        raise ValueError(f"Error al exportar el balance: {str(e)}")